import random
import pandas as pd
import numpy as np
from openpyxl import load_workbook

# se establece una seed (en cualquier numero) para que los datos sean consistentes entre aplicaciones
random.seed(2024)

# Hago la clase tractor para diferenciar los distintos mojamientos con su rendimiento diario
class Maquinaria:
    def __init__(self, rendimiento_c, rendimiento, tractor_activo, tractor_activo2, cosechadora_activa, name):

        self.rendimiento_c = rendimiento_c              # rendimiento de cosechadra
        self.rendimiento = rendimiento                  # rendimiento de tractor
        self.tractor_activo = tractor_activo            # arriendo de 1 tractor
        self.tractor_activo2 = tractor_activo2          # arriendo de 2 tractores
        self.cosechadora_activa = cosechadora_activa    # uso de cosechadora
        self.name = name                                # nombre de la aplicación

aplicacion1 = Maquinaria(13, 6, False, False, False, "300 litros por hectarea")
aplicacion2 = Maquinaria(12, 5, False,False,False, "400 litros por hectarea")
aplicacion3 = Maquinaria(11, 4.5, False,False,False, "500 litros por hectarea")

# asigno en valores fijos la cantidad de tractores y cosechadoras para uso dentro de la simulacion
tractores = 3 
cosechadora = 1

# calculo la probabilidad de que fallen 0, 1, 2 o 3 tractores

pf = 0.05       # probabilidad de que falle
npf = 0.95      # probabilidad de no falle

# Por triangulo de pascal, existen 4 escenarios:

p_0 =1*      npf**3     # ninguno falla
p_1 =3*pf*   npf**2     # 1 falla
p_2 =3*pf**2*npf        # 2 fallan
p_3 =1*pf**3            # 3 fallan

# redondeo a milesimas las probabilidades resultantes
p_0 = round(p_0, 3) 
p_1 = round(p_1, 3)
p_2 = round(p_2, 3)
p_3 = round(p_3, 3) 

# y cargo a un array de 4 opciones las 4 probabilidades de que ocurra tal escenario (como un dado cargado)
probabilidades = [p_0, p_1, p_2, p_3]   # el dado trabaja con una precicion de 0.999 debido a la perdida decimales

# realizo el mismo proceso para la cosechadora

npc = 0.9       #probabilidad de no fallo de la cosechadora
pc = 0.1        #probabilidad de fallo de la cosechadora

# en este caso el array tiene solo 2 datos (como una moneda cargada)
probabilidades_c = [npc, pc]

# creamos el espacio a simular
def simulacion(app):
    # definimos la variable de dias totales que tomo el proceso y hectareas acumuladas por dia

    aplicacion = app                    # carga de aplicacion a la funcion simulacion()
    dias_totales = 0                    # contador de dias 
    hectareas_totales = 0               # contador de hectareas
    hectareas = 220                     # total de hectareas por completar
    if app.tractor_activo == True:      # total de hectareas por completar con 1 tractor arrendado
        hectareas = 190
    elif app.tractor_activo2 == True:   # total de hectareas por completar con 2 tractores arrendado
        hectareas = 160
    cosechadora_count = 0

    # se crea un loop lo suficientemente grande para calcular los dias a trabajar
    for i in range(100):

### si el total acumulativo de hectareas es menor a las hectareas totales 
        if hectareas_totales<hectareas:

            ## Inicio un dia

            # se selecciona uno de los 4 escenarios de manera aleatoria y lo restamos de los tractores utilizados
            # esto para calcular los tractores operativos durante el dia
            escenario = random.choices([0, 1, 2, 3], weights=probabilidades, k=1) 
            tractores_totales = tractores - escenario[0]
            
            # para la cosechadora solo es necesario saber si falló o no
            escenario_c = random.choices([0, 1], weights=probabilidades_c, k=1) 

    ### Para trabajar los tractores:

            # si por lo menos uno falla
            if escenario[0] > 0:

                # iteramos este proceso por la cantidad de tractores con falla 
                rango_trabajado = 0 
                
                for i in range(escenario[0]):
                    # cargamos una sumatoria de un numero aleatorio entre 0 y el rendimiento en hectareas
                    # esto sera lo trabajado por los tractores antes de fallar
                    rango_trabajado = rango_trabajado + random.uniform(0,aplicacion.rendimiento)

                # finalmente calculamos el trabajo diario por los tractores con la siguiente formula:
                # rendimiento diario (en hectareas) * tractores operativos + rango trabajado por los tractores con falla
                dia_trabajado = aplicacion.rendimiento*(tractores_totales) + rango_trabajado
            
            # si ningun tractor falla
            else:
                # calculamos el trabajo diario multiplicando el rendimiento diario por la cantidad de tractores
                dia_trabajado = aplicacion.rendimiento*(tractores)

            # sumamos las hectareas trabajadas del dia al total de hectareas y sumamos un dia al total de dias
            hectareas_totales = hectareas_totales + dia_trabajado

    ### Para trabajar con la cosechadora:

            # si tenemos la cosechadora activa realizamos el mismo proceso que realizamos con los tractores
            if app.cosechadora_activa == True:

                # verificamos que la cosechadora no se encuentre en un dia de penalización
                # son 2 dias de reparación a la cosechadora luego de fallar
                if cosechadora_count == 0:

                    # si falla la cosechadora 
                    if escenario_c[0] == 1:      

                        # cargamos el dia trabajado con un numero entre 0 y el rendimiento en hectareas
                        # esto será lo trabajado por la cosechadora antes de fallar
                        dia_trabajado = random.uniform(0,aplicacion.rendimiento_c)

                        # como falló la cosechadora le damos 2 dias de penalización por falla
                        cosechadora_count = cosechadora_count + 2
                        
                    # si no falla la cosechadora
                    else:
                        # se carga el dia trabajado con el rendimiento total de la cosechadora
                        dia_trabajado = aplicacion.rendimiento_c   
                # si se encuentra en un dia de penalización, solo se descuenta uno de estos dias
                else:
                    
                    cosechadora_count = cosechadora_count - 1
                    pass

                # sumamos las hectareas trabajadas del dia al total de hectareas y sumamos un dia al total de dias
                hectareas_totales = hectareas_totales + dia_trabajado

            dias_totales = dias_totales+1

            # Fin del dia

### si el total acumulativo de hectareas es mayor o igual a las hectareas totales
        else:
            round(dias_totales, 3)
            break

    return dias_totales


# creamos un array donde se alojaran todos los resultados de las simulaciones
resultados = []

# escogemos la simulacion a trabajar
sim_app = aplicacion1

# ejecutamos la simulacion 1000 veces con la aplicacion preferida 
for i in range(100000):
    resultados.append(simulacion(sim_app))
# cargamos el array a un array de numpy para futuros calculos y graficos
resultados_np = np.array(resultados)

# imprimimos los 100 primeros resultados para verificar funcionamiento
print(f"\n_____Aplicacion utilizada: {sim_app.name}______\n\nVariables utilizadas\n  -Un tractor arrendado: {sim_app.tractor_activo}\n  -Dos tractor arrendados: {sim_app.tractor_activo2}\n  -Cosechadora utilizada: {sim_app.cosechadora_activa}")
print(f"\nMuestra de los resultados\n{resultados_np[:100]}\n")

serie = pd.Series(resultados_np, name="Valores")
# Descripción estadística
descripcion = serie.describe()
# Frecuencias de cada valor
frecuencias = serie.value_counts()

# Mostrar resultados
print(f"Descripción estadística básica:\n{descripcion}\n")

print(f"Frecuencias de cada valor:\n{frecuencias}")


"""
### Poblar excel con los resultados

# extraemos los valores específicos a poblar
datos = {
    "Promedio": descripcion["mean"],
    "Desviación": descripcion["std"],
    "Mínimo": descripcion["min"],
    "Máximo": descripcion["max"]
}

# cargamos la ruta del archivo de excel
ruta_excel = "Resultados.xlsx"

try:
    libro = load_workbook(ruta_excel)
except FileNotFoundError:
    # si el archivo no existe, creamos uno nuevo
    libro = load_workbook()

# cargamos la hoja a trabajar
nombre_hoja = "Tabla"

# verificamos si la hoja existe, creamos la hoja de no existir
if nombre_hoja not in libro.sheetnames:
    hoja = libro.create_sheet(title=nombre_hoja)
else:
    hoja = libro[nombre_hoja]

# fila a poblar
fila = 3

# cargamos las ubicaciones dentro de la fila para cada dato
ubicaciones = {
    "Promedio": f"D{fila}",
    "Desviación": f"E{fila}",
    "Mínimo": f"F{fila}",
    "Máximo": f"G{fila}"
}

# Poblamos las celdas con los valores
for key, valor in datos.items():
    celda = ubicaciones[key]
    hoja[celda] = valor

# guardamos los cambios
libro.save(ruta_excel)
"""

